import openpyxl

wb = openpyxl.load_workbook('planilla.xlsx')
sheetnames = wb.sheetnames


for sheet in sheetnames:
    print(wb[sheet])

wb.create_sheet('Hoja3')
print(wb.sheetnames)

hoja = wb.active
celda = hoja['A1']
celda.value = 'Nombre completo'
print(celda.value)

columna = hoja['A']
print(columna)

for col in columna:
    print(col.value)

wb.save('planilla_actualizado.xlsx')
